"""
PERFECT-Python — Interactive Launcher
======================================
Run this script from your perfect/ folder:

    python launch.py

It will:
  1. Ask for your data folder (where your .soil, .vege, .xlsx files live)
  2. Show you what it found and let you pick inputs
  3. Ask for a SILO station (name search or lat,lon)
  4. Ask for a date range
  5. Run the simulation and save outputs

No editing of file paths needed.
"""

import sys, os, json
from pathlib import Path

# ── make sure we can import our modules ───────────────────────────────────
# Use resolve() so __file__ is always absolute, and also insert cwd.
# This ensures cover_excel, soil, vege etc. are findable even when
# run_simulation.py is imported as a sub-module and resolves its own __file__
# to a different location.
HERE = Path(__file__).resolve().parent
for _p in [str(HERE), str(Path.cwd().resolve())]:
    if _p not in sys.path:
        sys.path.insert(0, _p)


# ── terminal helpers ───────────────────────────────────────────────────────

def header(text):
    print()
    print("─" * 60)
    print(f"  {text}")
    print("─" * 60)

def ask(prompt, default=None):
    if default:
        response = input(f"  {prompt} [{default}]: ").strip()
        return response if response else default
    else:
        while True:
            response = input(f"  {prompt}: ").strip()
            if response:
                return response
            print("  (required — please enter a value)")

def pick(prompt, options, allow_skip=False):
    """Show a numbered list and return the chosen item."""
    print(f"\n  {prompt}")
    for i, opt in enumerate(options, 1):
        print(f"    {i:2d}.  {opt}")
    if allow_skip:
        print(f"     0.  Skip / none")
    while True:
        raw = input("  Enter number: ").strip()
        if allow_skip and raw == '0':
            return None
        try:
            n = int(raw)
            if 1 <= n <= len(options):
                return options[n - 1]
        except ValueError:
            pass
        print(f"  Please enter a number between {'0' if allow_skip else '1'} and {len(options)}")

def confirm(prompt):
    r = input(f"  {prompt} [y/n]: ").strip().lower()
    return r in ('y', 'yes', '')


# ── scan data folder ───────────────────────────────────────────────────────


def _is_soil_excel(filepath):
    """
    Sniff an Excel file to decide if it is a soil description (not a cover schedule).
    Checks the first 6 rows of column A for soil-specific keywords.
    This means any .xlsx can be a soil file — no special naming needed.
    """
    SOIL_KEYWORDS = ('soil name', 'number of horizon', 'layer depth',
                     'wilting point', 'field capacity', 'number of layers')
    COVER_KEYWORDS = ('day/month', 'day no', 'green cover', 'cover model',
                      'data preparation')
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        col_a = []
        for row in ws.iter_rows(max_row=6, max_col=1, values_only=True):
            val = str(row[0] or '').strip().lower()
            if val:
                col_a.append(val)
        wb.close()
        for val in col_a:
            if any(kw in val for kw in COVER_KEYWORDS):
                return False
        for val in col_a:
            if any(kw in val for kw in SOIL_KEYWORDS):
                return True
    except Exception:
        pass
    # Filename fallback if content sniff inconclusive
    name_lower = str(filepath.stem).lower()
    return any(kw in name_lower for kw in ('soil', 'profile', 'horizon'))


def scan_folder(folder):
    folder = Path(folder)
    found = {
        'soil': [],       # .soil (XML), .PRM, or soil .xlsx
        'vege': [],       # .vege
        'excel': [],      # .xlsx cover schedules
        'soil_excel': [], # .xlsx soil description files
        'met': [],        # .MET climate files
        'p51': [],        # SILO P51 files
    }
    for f in sorted(folder.rglob('*')):
        if not f.is_file():
            continue
        ext = f.suffix.lower()
        if ext == '.soil':
            found['soil'].append(f)
        elif ext == '.prm':
            found['soil'].append(f)
        elif ext == '.vege':
            found['vege'].append(f)
        elif ext in ('.xlsx', '.xls'):
            if _is_soil_excel(f):
                found['soil_excel'].append(f)
                found['soil'].append(f)   # appears in soil picker
            else:
                found['excel'].append(f)
        elif ext == '.met':
            found['met'].append(f)
        elif ext == '.p51':
            found['p51'].append(f)
    return found


# ── SILO station search ────────────────────────────────────────────────────

def search_silo_stations(query, email):
    """Search SILO for stations matching a name fragment."""
    import requests
    url = 'https://www.longpaddock.qld.gov.au/cgi-bin/silo/PatchedPointDataset.php'
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (compatible; PERFECT-Python/1.0)',
            'Accept': 'text/plain, */*',
        }
        resp = requests.get(url, params={'format': 'name', 'nameFrag': query},
                            headers=headers, timeout=15)
        lines = [l.strip() for l in resp.text.strip().split('\n')
                 if l.strip() and '|' in l]
        stations = []
        for line in lines[:20]:
            parts = [p.strip() for p in line.split('|')]
            if len(parts) < 4:
                continue
            # Skip header rows (lat/lon fields are not numeric)
            try:
                lat_val = float(parts[2])
                lon_val = float(parts[3])
            except (ValueError, IndexError):
                continue
            # Skip obviously invalid coordinates
            if not (-45 < lat_val < -10 and 110 < lon_val < 155):
                continue
            stations.append({
                'number': parts[0],
                'name':   parts[1],
                'lat':    lat_val,
                'lon':    lon_val,
                'state':  parts[4] if len(parts) > 4 else '',
            })
        return stations
    except Exception as e:
        print(f"  Warning: could not reach SILO ({e})")
        return []


# ── main ───────────────────────────────────────────────────────────────────

def main():
    print()
    print("╔══════════════════════════════════════════════════════════╗")
    print("║         PERFECT-Python  ·  Water Balance Model          ║")
    print("╚══════════════════════════════════════════════════════════╝")

    # ── 1. Data folder ─────────────────────────────────────────────────────
    header("Step 1 — Data folder")
    print("  Enter the path to the folder containing your input files.")
    print("  This will be scanned for .soil, .PRM, .vege and .xlsx files.")
    print()

    while True:
        folder_str = ask("Data folder path", default=str(Path.cwd()))
        folder = Path(folder_str).expanduser().resolve()
        if folder.exists():
            break
        print(f"  ✗ Folder not found: {folder}")

    print(f"\n  Scanning {folder} ...")
    found = scan_folder(folder)

    soil_files  = found['soil']
    vege_files  = found['vege']
    excel_files = found['excel']
    met_files   = found['met']
    p51_files   = found['p51']

    total = sum(len(v) for v in found.values())
    soil_excel_files = found.get('soil_excel', [])
    print(f"  Found: {len(soil_files)} soil  ·  {len(vege_files)} vege  ·  "
          f"{len(excel_files)} cover xlsx  ·  {len(soil_excel_files)} soil xlsx  ·  "
          f"{len(met_files)} .MET  ·  {len(p51_files)} .P51")

    if total == 0:
        print("\n  No recognised input files found in that folder.")
        print("  Make sure your .soil, .PRM, .vege or .xlsx files are there.")
        sys.exit(1)

    # ── 2. Soil ────────────────────────────────────────────────────────────
    header("Step 2 — Soil profile")

    if not soil_files:
        print("  No soil files (.soil or .PRM) found in that folder.")
        sys.exit(1)

    soil_labels = [f.name for f in soil_files]
    soil_choice = pick("Select soil file:", soil_labels)
    soil_path   = soil_files[soil_labels.index(soil_choice)]
    print(f"\n  Selected: {soil_path}")

    # Quick parse to show PAWC
    try:
        ext = soil_path.suffix.lower()
        if ext == '.soil':
            from soil_xml import read_soil_xml
            profile = read_soil_xml(soil_path)
        elif ext in ('.xlsx', '.xls'):
            from soil_excel import read_soil_excel
            profile = read_soil_excel(soil_path)
        else:
            from soil import read_prm
            profile = read_prm(soil_path)
        print(f"  → {profile.name}  ·  {len(profile.layers)} layers  ·  "
              f"PAWC = {profile.pawc_total:.0f} mm  ·  "
              f"CN2 = {profile.cn2_bare}  ·  Cona = {profile.cona}")
    except Exception as e:
        print(f"  Warning: could not parse soil file ({e})")
        profile = None

    # ── 3. Vegetation ──────────────────────────────────────────────────────
    header("Step 3 — Vegetation / crop")

    all_vege = vege_files + excel_files
    vege_labels = [f.name for f in all_vege]

    if not all_vege:
        print("  No vegetation files found (.vege or .xlsx).")
        use_bare = confirm("  Run as bare fallow?")
        if not use_bare:
            sys.exit(1)
        vege_path = None
        vege_internal_name = 'bare_fallow'
    else:
        # Add bare fallow option
        vege_labels_display = vege_labels + ['[ Bare fallow — no crop ]']
        vege_choice = pick("Select vegetation file:", vege_labels_display)

        if vege_choice == '[ Bare fallow — no crop ]':
            vege_path = None
            print("\n  Selected: bare fallow")
        else:
            vege_path = all_vege[vege_labels.index(vege_choice)]
            print(f"\n  Selected: {vege_path}")

            # Quick parse to show cover summary
            try:
                ext = vege_path.suffix.lower()
                if ext == '.vege':
                    from vege import read_vege, get_vege_state
                    import numpy as np
                    v = read_vege(vege_path)
                    vege_internal_name = v.name if isinstance(v.name, str) else ''.join(v.name)
                    peak_g = max(get_vege_state(v,d)[0]*100 for d in range(1,366))
                    peak_r = max(get_vege_state(v,d)[2] for d in range(1,366))
                    print(f"  → {vege_internal_name}  ·  Peak green cover = {peak_g:.0f}%  ·  "
                          f"Max root depth = {peak_r:.0f} mm")
                else:
                    from cover_excel import read_cover_excel
                    import numpy as np
                    v = read_cover_excel(vege_path)
                    _raw_vname = v.name if isinstance(v.name, str) else ''.join(v.name)
                    vege_internal_name = _short(_raw_vname, VEGE_DROP)
                    peak_g = max(v.green_cover) * 100
                    peak_r = max(v.root_depth)
                    print(f"  → {vege_internal_name}  ·  Peak green cover = {peak_g:.0f}%  ·  "
                          f"Max root depth = {peak_r:.0f} mm")
            except Exception as e:
                print(f"  Warning: could not parse vege file ({e})")

    # ── 4. Climate ─────────────────────────────────────────────────────────
    header("Step 4 — Climate data")

    # Check for local climate files first (.P51 and .MET)
    use_silo = True
    station_info = {}
    met_path = None

    local_climate = p51_files + met_files
    if local_climate:
        local_labels = [f.name for f in local_climate] + ['[ Fetch from SILO instead ]']
        type_hints   = ['P51']*len(p51_files) + ['MET']*len(met_files) + ['']
        print(f"  Found {len(p51_files)} .P51 and {len(met_files)} .MET file(s) in your data folder.")
        met_choice = pick("Select a local climate file or fetch from SILO:", local_labels)
        if met_choice != '[ Fetch from SILO instead ]':
            met_path = local_climate[local_labels[:-1].index(met_choice)]
            use_silo = False
            print(f"\n  Selected: {met_path.name}")

    if use_silo:
        print("  Fetch daily climate from SILO (requires internet).")
        print("  Options:")
        print("    Type a station name  e.g.  Dalby  or  Oakey")
        print("    Or paste lat,lon     e.g.  -27.28, 151.26")
        print("  If the name search fails, use lat,lon directly.")
        print()

        email = ask("Your email (SILO API username)", default="david.freebairn@gmail.com")

        while True:
            query = ask("Station name or lat,lon")

            # Check if lat,lon
            import re
            ll = re.match(r'^(-?\d+\.?\d*)\s*,\s*(-?\d+\.?\d*)$', query)
            if ll:
                lat, lon = float(ll.group(1)), float(ll.group(2))
                station_info = {
                    'name': f'Grid ({lat:.3f}, {lon:.3f})',
                    'lat': lat, 'lon': lon,
                    'source': 'DataDrill interpolated',
                }
                print(f"\n  → Grid point at ({lat:.4f}, {lon:.4f})")
                break
            else:
                print(f"\n  Searching SILO for '{query}' ...")
                stations = search_silo_stations(query, email)
                if not stations:
                    # Show first few lines of response to help debug
                    preview = resp.text.strip()[:300] if hasattr(resp,'text') else ''
                    if 'rejected' in preview.lower():
                        print("  SILO search was blocked by server — try lat,lon instead")
                        print("  e.g.  -27.28, 151.26  for Dalby")
                    elif preview:
                        print("  No stations found (or unexpected response).")
                        print(f"  Response preview: {preview[:150]}")
                        print("  Try a different name or use lat,lon directly")
                    else:
                        print("  No stations found — try a different name or use lat,lon")
                    continue
                labels = [f"{s['name']}  (#{s['number']}  {s['lat']:.3f}, {s['lon']:.3f}  {s['state']})"
                          for s in stations]
                choice = pick("Select station:", labels)
                try:
                    idx = labels.index(choice)
                    s = stations[idx]
                except (ValueError, IndexError):
                    # pick() returned the string directly; match by position
                    idx = 0
                    s = stations[0]
                station_info = {
                    'name'  : s['name'],
                    'number': s['number'],
                    'lat'   : s['lat'],
                    'lon'   : s['lon'],
                    'source': 'Patched Point',
                }
                print(f"\n  → {s['name']}  ({s['lat']:.4f}, {s['lon']:.4f})")
                break

    # ── Peek date bounds from local climate file ────────────────────────────
    # Used to prefill and validate Step 5 date range inputs
    _clim_start = None   # earliest date in file  (YYYYMMDD string)
    _clim_end   = None   # latest date in file     (YYYYMMDD string)
    if not use_silo and met_path is not None:
        try:
            ext = met_path.suffix.lower()
            if ext == '.p51':
                from read_p51 import read_p51 as _rp51
                _, _peek = _rp51(met_path)
            else:
                from perfect_io import read_met as _rmet
                _, _peek = _rmet(met_path)
            _clim_start = _peek.index.min().strftime('%Y%m%d')
            _clim_end   = _peek.index.max().strftime('%Y%m%d')
            print(f"\n  Climate file covers: {_clim_start} → {_clim_end}"
                  f"  ({_peek.index.year.nunique()} years)")
        except Exception as _e:
            print(f"  (Could not read date range from file: {_e})")

    # ── 5. Date range ──────────────────────────────────────────────────────
    header("Step 5 — Simulation period")

    _default_start = _clim_start or "19570101"
    _default_end   = _clim_end   or "19981231"

    if _clim_start:
        print(f"  Climate data available: {_clim_start} → {_clim_end}")
        print(f"  Dates must be within this range.")
        print()

    while True:
        start = ask("Start date (YYYYMMDD)", default=_default_start)
        end   = ask("End date   (YYYYMMDD)", default=_default_end)

        # Basic format check
        import re as _re
        if not (_re.fullmatch(r'\d{8}', start) and _re.fullmatch(r'\d{8}', end)):
            print("  ✗ Dates must be 8 digits in YYYYMMDD format — please try again.")
            continue

        # Logical order check
        if start >= end:
            print(f"  ✗ Start date ({start}) must be before end date ({end}) — please try again.")
            continue

        # Bounds check against climate file
        if _clim_start:
            if start < _clim_start:
                print(f"  ✗ Start date {start} is before the first record in the climate file "
                      f"({_clim_start}) — please try again.")
                continue
            if end > _clim_end:
                print(f"  ✗ End date {end} is after the last record in the climate file "
                      f"({_clim_end}) — please try again.")
                continue

        break   # all checks passed

    start_yr = start[:4]; end_yr = end[:4]
    nyears = int(end_yr) - int(start_yr) + 1
    print(f"\n  → {start_yr} to {end_yr}  ({nyears} years)")

    # ── 6. Output folder ───────────────────────────────────────────────────
    header("Step 6 — Output")

    out_folder_str = ask("Output folder", default=str(folder / 'results'))
    out_folder = Path(out_folder_str).expanduser().resolve()
    out_folder.mkdir(parents=True, exist_ok=True)

    # Suggest a run name
    def _short(stem, drop_words):
        """Remove common redundant words from a file stem and clean up."""
        import re
        s = stem
        for w in drop_words:
            s = re.sub(w, '', s, flags=re.IGNORECASE)
        s = re.sub(r'[_\-]+', '_', s).strip('_')   # collapse repeated separators
        return s or stem   # fall back to original if we stripped everything

    SOIL_DROP = [
        r'[_\-]?[Ss]oil[_\-]?[Pp]rofile[_\-]?',
        r'[_\-]?[Pp]rofile[_\-]?',
    ]
    VEGE_DROP = [
        r'[_\-]?[Vv]egetation[_\-]?[Dd]escription[_\-]?',
        r'[_\-]?[Cc]over[_\-]?[Dd]ata[_\-]?[Ff]or[_\-]?[Hh]owleaky[_\-]?',
        r'[_\-]?[Dd]ata[_\-]?[Pp]reparation[_\-]?[Ff]or[_\-]?[Hh]owleaky[_\-]?',
        r'[_\-]?[Cc]over[_\-]?[Dd]ata[_\-]?',
        r'[_\-]?[Cc]rop[_\-]?[Dd]ata[_\-]?',
        r'[_\-]?[Dd]escription[_\-]?',
        r'[_\-]?[Ss]chedule[_\-]?',
    ]

    import re as _re2
    def _fn(s, n=35):
        return _re2.sub(r'[\s_\-]+', '_', str(s)).strip('_')[:n]
    # Use internal names read from the files, not the filenames
    soil_label = _fn(profile.name) if profile else _short(soil_path.stem, SOIL_DROP) if soil_path else 'unknown'
    vege_label_fn = _fn(locals().get('vege_internal_name', '') or
                        (vege_path.stem if vege_path else 'bare_fallow'))
    if use_silo:
        stn_label_fn = _fn(station_info.get('name', 'station'))
    else:
        stn_label_fn = _fn(met_path.stem if met_path else 'climate')
    default_name = f"{stn_label_fn}_{soil_label}_{vege_label_fn}"
    run_name = ask("Run name (used for output filenames)", default=default_name)

    # Append datetime stamp so every run produces uniquely named files
    run_name_dt = run_name   # no datetime stamp — keep filenames short

    # ── 7. Summary + confirm ───────────────────────────────────────────────
    header("Ready to run — summary")
    if use_silo:
        print(f"  Climate  : {station_info['name']}  ({station_info['source']})")
        print(f"             lat={station_info['lat']:.4f}  lon={station_info['lon']:.4f}")
    else:
        ftype = "SILO P51" if met_path.suffix.lower()==".p51" else "PERFECT .MET"
        print(f"  Climate  : {met_path.name}  ({ftype})")
    print(f"  Soil     : {soil_path.name}")
    print(f"  Vege     : {vege_path.name if vege_path else 'Bare fallow'}")
    print(f"  Period   : {start} → {end}  ({nyears} years)")
    print(f"  Outputs  : {out_folder}")
    print()

    if not confirm("Run simulation?"):
        print("\n  Cancelled.")
        sys.exit(0)

    # ── 8. Run ─────────────────────────────────────────────────────────────
    header("Running simulation")

    # Build config for run_simulation.run_from_config()
    # We handle climate separately here so we can show progress

    if use_silo:
        from silo_fetch import fetch_silo
        cache_path = out_folder / f"silo_{station_info['lat']:.3f}_{station_info['lon']:.3f}_{start}_{end}.csv"
        # Ensure cache directory exists (out_folder may not have been written yet)
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        print(f"  Fetching SILO climate data ...")
        try:
            _, met_df = fetch_silo(
                lat=station_info['lat'], lon=station_info['lon'],
                start=start, end=end,
                email=email,
                cache_path=str(cache_path),
            )
        except Exception as fetch_err:
            print(f"\n  SILO fetch failed: {fetch_err}")
            print("  Tip: if you see 'Request Rejected', your network is blocking SILO.")
            print("  Download the .P51 file from https://www.longpaddock.qld.gov.au/silo/")
            print("  and re-run choosing the local file option in Step 4.")
            raise
    else:
        ext = met_path.suffix.lower()
        if ext == '.p51':
            from read_p51 import read_p51
            print(f"  Loading {met_path.name} (SILO P51) ...")
            _, met_df = read_p51(met_path)
        else:
            from perfect_io import read_met
            print(f"  Loading {met_path.name} (PERFECT .MET) ...")
            _, met_df = read_met(met_path)
        # Trim to requested date range if user specified a sub-period
        met_df = met_df[start:end]

    print(f"  {len(met_df)} days loaded  ({met_df.index[0].date()} → {met_df.index[-1].date()})")

    # Load soil
    ext = soil_path.suffix.lower()
    if ext == '.soil':
        from soil_xml import read_soil_xml
        profile = read_soil_xml(soil_path)
    elif ext in ('.xlsx', '.xls'):
        from soil_excel import read_soil_excel
        profile = read_soil_excel(soil_path)
    else:
        from soil import read_prm
        profile = read_prm(soil_path)

    # Load vege
    if vege_path is None:
        get_state = lambda doy: (0.0, 0.0, 0.0)
    else:
        ext = vege_path.suffix.lower()
        if ext == '.vege':
            from vege import read_vege
            from run_simulation import _make_vege_fn
            get_state = _make_vege_fn(read_vege(vege_path))
        else:
            from cover_excel import read_cover_excel
            from run_simulation import _make_cover_fn
            get_state = _make_cover_fn(read_cover_excel(vege_path))

    # Run daily model
    from run_simulation import _run_daily, _monthly_means, _annual_stats
    print(f"  Running daily water balance ...")
    df, sw0, sw_final = _run_daily(met_df, profile, get_state)
    nyears_actual = met_df.index.year.nunique()
    dsw = sw_final - sw0
    err = (df.rain.sum() - df.runoff.sum() - df.drainage.sum()
           - df.soil_evap.sum() - df.transp.sum() - dsw)

    print(f"  Done.  Balance error = {err:.4f} mm  ✓" if abs(err) < 0.01
          else f"  Done.  Balance error = {err:.4f} mm  ⚠ CHECK")
    err_val = err / nyears_actual
    swf     = sw_final

    mon = _monthly_means(df, nyears_actual)
    ann = _annual_stats(df)
    ann.attrs['annual_yield'] = df.attrs.get('annual_yield', {})

    # ── 9. Print results ───────────────────────────────────────────────────
    header("Results")
    print(f"  {'Component':<22}  {'Mean mm/yr':>10}  {'% of rain':>10}  {'CV%':>6}")
    print(f"  {'─'*54}")
    rain_mean = float(ann.rain.mean())
    for k, label in [('rain','Rainfall'),('runoff','Runoff'),
                     ('drainage','Deep drainage'),('soil_evap','Soil evap'),
                     ('transp','Transpiration'),('et','Total ET')]:
        v = float(ann[k].mean())
        cv = float(ann[k].std() / max(ann[k].mean(), 0.1) * 100)
        pct = v / rain_mean * 100
        print(f"  {label:<22}  {v:>10.1f}  {pct:>9.1f}%  {cv:>5.0f}%")
    print(f"\n  Delta SW  : {dsw:.1f} mm total  ({dsw/nyears_actual:.2f} mm/yr)")
    print(f"  Balance   : {err:.4f} mm")
    yd = df.attrs.get("annual_yield", {})
    if yd:
        mean_y = sum(yd.values()) / len(yd)
        print(f"  Yield     : {mean_y:.2f} t/ha/yr mean  ({len(yd)} seasons)")
        for yr, yv in sorted(yd.items()):
            print(f"              {yr}: {yv:.2f} t/ha")

    # ── 10. Save outputs ───────────────────────────────────────────────────
    header("Saving outputs")
    import numpy as np, pandas as pd

    # CSV — annual
    ann_out = ann[['rain','runoff','drainage','soil_evap','transp','et']].copy()
    ann_out.columns = ['rain_mm','runoff_mm','drainage_mm','soil_evap_mm','transp_mm','et_mm']
    # Add annual yield if available
    yd = df.attrs.get('annual_yield', {})
    if yd:
        ann_out['yield_t_ha'] = ann_out.index.map(lambda y: round(yd.get(y, 0.0), 3))
    csv_path = out_folder / f"{run_name_dt}_annual.csv"
    ann_out.to_csv(csv_path)
    print(f"  Saved: {csv_path.name}")

    # CSV — monthly means
    mon_out = mon[['rain','runoff','drainage','soil_evap','transp','et','dsw']].copy()
    mon_out.index.name = 'month'
    mon_csv = out_folder / f"{run_name_dt}_monthly.csv"
    mon_out.to_csv(mon_csv)
    print(f"  Saved: {mon_csv.name}")

    # CSV — daily output
    from run_simulation import save_daily_csv
    daily_csv = out_folder / f"{run_name_dt}_daily.csv"
    save_daily_csv(df, daily_csv)

    # JSON — full results
    results_dict = {
        'meta': {
            'run_name': run_name_dt,
            'soil': str(soil_path),
            'vege': str(vege_path) if vege_path else 'bare fallow',
            'climate': station_info if use_silo else str(met_path),
            'start': start, 'end': end, 'nyears': nyears_actual,
            'pawc_mm': round(profile.pawc_total, 1),
            'balance_error_mm': round(err, 4),
        },
        'annual_means': {
            k: round(float(ann[k].mean()), 1)
            for k in ['rain','runoff','drainage','soil_evap','transp','et']
        },
        'monthly': {
            k: mon[k].round(1).tolist()
            for k in ['rain','runoff','drainage','soil_evap','transp','et','dsw']
        },
    }
    json_path = out_folder / f"{run_name_dt}_results.json"
    with open(json_path, 'w') as f:
        json.dump(results_dict, f, indent=2)
    print(f"  Saved: {json_path.name}")

    # PNG — Run summary chart (full layout)
    try:
        import sys as _sys
        _sys.path.insert(0, str(__import__('pathlib').Path(__file__).parent))
        from output_chart import make_output_chart
        stn_label  = station_info.get('name', '') if use_silo else met_path.stem
        vege_label = vege_path.stem if vege_path else 'Bare fallow'
        title_str  = (f'{stn_label}  ·  {profile.name}  ·  {vege_label}  ·  '
                      f'{start[:4]}–{end[:4]}  ({nyears_actual} years)  ·  '
                      f'Balance error = {err_val:+.3f} mm/yr')
        png_path = out_folder / f"{run_name_dt}_Run_summary.png"
        vege_label_full = vege_path.stem if vege_path else 'Bare fallow'
        stn_full = station_info.get('name','') if use_silo else met_path.stem
        make_output_chart(ann, mon, profile, nyears_actual,
                          swf - sw0, err_val,
                          title_str=title_str,
                          out_path=png_path,
                          climate_name=stn_full,
                          soil_name=profile.name,
                          crop_name=vege_label_full)
    except Exception as e:
        import traceback
        print(f"  Chart not saved: {e}")
        traceback.print_exc()

    # Input summary graphics
    try:
        from input_summaries import make_soil_summary, make_vege_summary
        make_soil_summary(profile,
            out_folder / f"{run_name_dt}_{profile.name.replace(' ','_')}_soil_summary.png")
        if vege_path is not None:
            ext_v = vege_path.suffix.lower()
            if ext_v == '.vege':
                from vege import read_vege
                _vobj = read_vege(vege_path)
            else:
                from cover_excel import read_cover_excel
                _vobj = read_cover_excel(vege_path)
                if _vobj.tue > 0 and _vobj.hi > 0:
                    print(f"  TUE={_vobj.tue:.1f} g/m²/mm  HI={_vobj.hi:.2f} — yield will be estimated")
                else:
                    print("  No TUE/HI in Excel file — yield will not be estimated")
            make_vege_summary(_vobj,
                out_folder / f"{run_name_dt}_{vege_path.stem}_vege_summary.png")
    except Exception as _e:
        print(f"  Input summaries not saved: {_e}")

    # Monthly summary chart (HowLeaky style)
    try:
        from monthly_chart import make_monthly_chart
        mon_png = out_folder / f"{run_name_dt}_monthly_summary.png"
        stn_lbl = station_info.get('name','') if use_silo else met_path.stem
        vege_lbl = vege_path.stem if vege_path else 'Bare fallow'
        yr_range = f"{start[:4]}–{end[:4]}"
        make_monthly_chart(
            ann, mon, profile, nyears_actual, err_val,
            title_str    = f'Water balance summary  {stn_lbl}  {yr_range}',
            out_path     = mon_png,
            climate_name = stn_lbl,
            soil_name    = profile.name,
            crop_name    = vege_lbl,
        )
    except Exception as _e:
        print(f"  Monthly chart not saved: {_e}")

    header("Complete")
    print(f"  All outputs saved to:  {out_folder}")
    print(f"  Run again anytime:     python launch.py")
    print()


if __name__ == '__main__':
    main()
