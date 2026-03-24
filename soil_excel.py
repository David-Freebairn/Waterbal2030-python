"""
soil_excel.py
=============
Reads a soil profile from the 'Soil description' Excel format and returns
a SoilProfile object identical to what soil_xml.py produces — so it drops
straight into run_simulation.py with no other changes.

Expected Excel layout (single sheet, row-labelled):
  Row 1  : Soil name         | <name>
  Row 3  : Number of Horizons| <n>
  Row 4  : Layer Depth       | val1 | val2 | ... | mm
  Row 5  : Air dry moisture  | val1 | ...        | %Vol
  Row 6  : Wilting point     | val1 | ...        | %Vol
  Row 7  : Field capacity    | val1 | ...        | %Vol
  Row 8  : Sat. water content| val1 | ...        | %Vol
  Row 9  : Max. drainage     | val1 | ...        | mm/day
  Row 10 : Bulk density      | val1 | ...        | g/cm3
  Row 12 : Stage 1 evap (U)  | val
  Row 13 : Stage 2 evap (Cona)| val
  Row 14 : Runoff Curve No.  | val
  Row 15 : CN reduction cover| val
  Row 16 : Erodibility (K)   | val
  Row 17 : Field Slope (S)   | val  (%)
  Row 18 : Slope Length (L)  | val  (m)
  Row 19 : Practice factor   | val
  Row 20 : CN Reduction Till | val
  Row 21 : Rainfall to 0 rough| val
  Row 23 : Rill/interrill    | val

Usage
-----
    from soil_excel import read_soil_excel
    profile = read_soil_excel('Soil_description.xlsx')

    # Works identically to soil_xml.py output:
    from run_simulation import _run_daily
    df, sw0, swf = _run_daily(met, profile, get_state)
"""

import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

# Import the dataclasses defined in soil.py
from soil import SoilProfile, SoilLayer


# ── Internal helpers ──────────────────────────────────────────────────────────

def _row_values(ws, label_fragment, n_values, default=0.0):
    """
    Find the row whose first cell contains label_fragment (case-insensitive),
    and return the next n_values numeric cells as a list of floats.
    """
    label_fragment = label_fragment.lower()
    for row in ws.iter_rows(values_only=True):
        cell0 = str(row[0] or '').lower()
        if label_fragment in cell0:
            vals = []
            for cell in row[1:]:
                if cell is None:
                    continue
                try:
                    v = float(cell)
                    vals.append(v)
                    if len(vals) == n_values:
                        break
                except (TypeError, ValueError):
                    continue
            # Pad with default if fewer values found than expected
            while len(vals) < n_values:
                vals.append(default)
            return vals
    return [default] * n_values


def _scalar(ws, label_fragment, default=0.0):
    """Return the first numeric value from the row matching label_fragment."""
    return _row_values(ws, label_fragment, 1, default)[0]


def _soil_name(ws):
    """Read the soil name from the first data cell of the 'Soil name' row."""
    for row in ws.iter_rows(values_only=True):
        cell0 = str(row[0] or '').lower()
        if 'soil name' in cell0:
            for cell in row[1:]:
                if cell is not None and str(cell).strip():
                    return str(cell).strip()
    return 'Unknown soil'


def _n_horizons(ws):
    """Read the number of horizons."""
    val = _scalar(ws, 'number of horizon', default=4)
    return max(1, int(val))


# ── Public API ────────────────────────────────────────────────────────────────

def read_soil_excel(filepath, sheet_name=None) -> SoilProfile:
    """
    Parse a Soil_description Excel file into a SoilProfile.

    Parameters
    ----------
    filepath   : str or Path — path to .xlsx file
    sheet_name : str or None — sheet to read; if None, uses the first sheet

    Returns
    -------
    SoilProfile  (same structure as soil_xml.read_soil_xml output)
    """
    filepath = Path(filepath)
    wb = load_workbook(filepath, read_only=True, data_only=True)

    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    # ── Identity ──────────────────────────────────────────────────────────
    name = _soil_name(ws)
    n    = _n_horizons(ws)

    # ── Layer arrays (%Vol → fraction, depth in mm) ───────────────────────
    depths   = _row_values(ws, 'layer depth',        n)   # mm cumulative
    airdry   = _row_values(ws, 'air dry',            n)   # %Vol
    ll       = _row_values(ws, 'wilting point',      n)   # %Vol
    dul      = _row_values(ws, 'field capacity',     n)   # %Vol
    sat      = _row_values(ws, 'sat. water',         n)   # %Vol
    drain_d  = _row_values(ws, 'max. drainage',      n)   # mm/day
    bulk_d   = _row_values(ws, 'bulk density',       n)   # g/cm3

    # ── Build SoilLayer objects ───────────────────────────────────────────
    layers = []
    prev_depth = 0.0
    for i in range(n):
        depth = float(depths[i])
        thick = depth - prev_depth

        ad = np.clip(airdry[i] / 100.0, 0.0, 1.0)
        l  = np.clip(ll[i]     / 100.0, 0.0, 1.0)
        d  = np.clip(dul[i]    / 100.0, 0.0, 1.0)
        s  = np.clip(sat[i]    / 100.0, 0.0, 1.0)

        # Max drainage rate (mm/day) → Ksat (mm/hr) — matches soil_xml.py
        ksat = drain_d[i] / 24.0

        layer = SoilLayer(
            depth_mm     = depth,
            thickness    = thick,
            airdry       = ad,
            ll           = l,
            dul          = d,
            sat          = s,
            ksat         = ksat,
            ll_mm        = l * thick,
            dul_mm       = d * thick,
            sat_mm       = s * thick,
            airdry_mm    = ad * thick,
            pawc         = (d - l) * thick,
            bulk_density = bulk_d[i],
        )
        layers.append(layer)
        prev_depth = depth

    # ── Scalar soil parameters ────────────────────────────────────────────
    u         = _scalar(ws, 'stage 1 evap',           default=9.0)
    cona      = _scalar(ws, 'stage 2 evap',           default=4.0)
    cn2       = _scalar(ws, 'runoff curve',           default=85.0)
    cn_cv     = _scalar(ws, 'cn reduction cover',     default=20.0)
    k_fac     = _scalar(ws, 'erodibility',            default=0.48)
    slope     = _scalar(ws, 'field slope',            default=3.0)
    sl        = _scalar(ws, 'slope length',           default=100.0)
    p_fac     = _scalar(ws, 'practice factor',        default=1.0)
    cn_tl     = _scalar(ws, 'cn reduction - till',    default=0.0)
    cn_rn     = _scalar(ws, 'rainfall to 0 rough',    default=0.0)
    rill      = _scalar(ws, 'rill',                   default=1.0)

    mean_bd   = float(np.mean(bulk_d))

    profile = SoilProfile(
        name                = name,
        layers              = layers,
        cona                = cona,
        u                   = u,
        cn2_bare            = cn2,
        cn_cover_reduction  = cn_cv,
        cn_tillage_max      = cn_tl,
        cn_roughness_rain   = cn_rn,
        musle_k             = k_fac,
        musle_p             = p_fac,
        slope_pct           = slope,
        slope_length        = sl,
        rill_ratio          = rill,
        bulk_density        = mean_bd,
        cracking            = False,   # not in this format — extend if needed
        crack_infil         = 10.0,
    )
    profile.total_depth = float(depths[-1])
    profile.pawc_total  = sum(l.pawc for l in layers)

    # ── Summary print (matches soil_xml.py style) ─────────────────────────
    print(f"  Soil   : {profile.name}")
    print(f"  Layers : {n}   Depth: {profile.total_depth:.0f} mm   "
          f"PAWC: {profile.pawc_total:.0f} mm")
    print(f"  CN2={profile.cn2_bare}  CN cover reduction={profile.cn_cover_reduction}  "
          f"Cona={profile.cona}  U={profile.u}")

    return profile


# ── Standalone test ───────────────────────────────────────────────────────────

if __name__ == '__main__':
    import sys
    fpath = sys.argv[1] if len(sys.argv) > 1 else '/mnt/user-data/uploads/Soil_description.xlsx'
    p = read_soil_excel(fpath)

    print()
    print(f"{'Lyr':>4}  {'Thick':>6}  {'AirDry':>7}  {'LL':>6}  "
          f"{'DUL':>6}  {'SAT':>6}  {'Ksat':>8}  {'PAWC':>6}")
    print("-" * 60)
    for i, l in enumerate(p.layers):
        print(f"  {i+1:2d}  {l.thickness:6.0f}  "
              f"{l.airdry*100:6.1f}%  {l.ll*100:5.1f}%  "
              f"{l.dul*100:5.1f}%  {l.sat*100:5.1f}%  "
              f"{l.ksat:6.2f}mm/hr  {l.pawc:5.1f}mm")
    print(f"\n  Total PAWC : {p.pawc_total:.0f} mm")
    print(f"  Total depth: {p.total_depth:.0f} mm")
